import os
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
PRIVATE_BASE = Path(os.getenv("WYZE_OUTPUT_DIR", PROJECT_DIR / "output"))
CLEAN_MASTER = PRIVATE_BASE / "clean" / "Wyze_Scale_Clean_Master.xlsx"
ANALYSIS_DIR = PRIVATE_BASE / "analysis"
OUTPUT_ANALYSIS = ANALYSIS_DIR / "Wyze_Scale_Trend_Analysis.xlsx"


BRAND_FILL = "006573"
LIGHT_FILL = "D9EAF0"
SOFT_FILL = "E2D7C2"
ORANGE_FILL = "F47D20"
WHITE = "FFFFFF"
BLACK = "000000"
BORDER_COLOR = "B7B7B7"


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor=BRAND_FILL)
    font = Font(bold=True, color=WHITE)
    border = Border(bottom=Side(style="thin", color=BORDER_COLOR))

    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws, max_width=28):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0

        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def write_df(ws, df, start_row=1, start_col=1):
    headers = list(df.columns)

    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=header)

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            if pd.isna(value):
                value = None
            ws.cell(row=row_idx, column=col_idx, value=value)

    style_header(ws, start_row)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col).coordinate
    autofit(ws)


def add_line_chart(ws, title, data_col, date_col=1, anchor="H2", height=9, width=18):
    max_row = ws.max_row
    chart = LineChart()
    chart.title = title
    chart.style = 13
    chart.y_axis.title = title
    chart.x_axis.title = "Date"
    chart.height = height
    chart.width = width

    data = Reference(ws, min_col=data_col, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=date_col, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, anchor)


def add_bar_chart(ws, title, data_col, category_col=1, anchor="H2", height=9, width=18):
    max_row = ws.max_row
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = title
    chart.x_axis.title = "Period"
    chart.height = height
    chart.width = width

    data = Reference(ws, min_col=data_col, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=category_col, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, anchor)


def safe_latest(series):
    clean = series.dropna()
    if clean.empty:
        return None
    return clean.iloc[-1]


def main():
    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)

    if not CLEAN_MASTER.exists():
        raise FileNotFoundError(f"Cannot find clean master workbook: {CLEAN_MASTER}")

    daily = pd.read_excel(CLEAN_MASTER, sheet_name="Daily_Analysis")
    analysis_ready = pd.read_excel(CLEAN_MASTER, sheet_name="Analysis_Ready")
    clean_all = pd.read_excel(CLEAN_MASTER, sheet_name="Clean_All")

    for df in [daily, analysis_ready, clean_all]:
        if "local_date" in df.columns:
            df["local_date"] = pd.to_datetime(df["local_date"])
        if "local_datetime" in df.columns:
            df["local_datetime"] = pd.to_datetime(df["local_datetime"])

    daily = daily.sort_values("local_date").reset_index(drop=True)
    analysis_ready = analysis_ready.sort_values("local_datetime").reset_index(drop=True)

    # Trend data
    trend_cols = [
        "local_date",
        "weight",
        "weight_kg",
        "bmi",
        "body_fat",
        "muscle",
        "body_water",
        "protein",
        "body_vfr",
    ]
    trend_cols = [c for c in trend_cols if c in daily.columns]
    trend = daily[trend_cols].copy()

    trend["weight_7_day_avg"] = trend["weight"].rolling(7, min_periods=1).mean()
    trend["weight_30_day_avg"] = trend["weight"].rolling(30, min_periods=1).mean()

    bodycomp_cols = [
        "local_date",
        "body_fat",
        "muscle",
        "body_water",
        "protein",
        "bone_mineral",
        "body_vfr",
    ]
    bodycomp_cols = [c for c in bodycomp_cols if c in daily.columns]
    bodycomp = daily[bodycomp_cols].copy()
    metric_cols = [c for c in bodycomp.columns if c != "local_date"]
    bodycomp = bodycomp.dropna(subset=metric_cols, how="all")

    # Monthly summary
    monthly = daily.copy()
    monthly["month"] = monthly["local_date"].dt.to_period("M").astype(str)

    agg_map = {
        "weight": ["mean", "min", "max", "count"],
        "bmi": "mean",
        "body_fat": "mean",
        "muscle": "mean",
        "body_water": "mean",
    }

    existing_agg = {k: v for k, v in agg_map.items() if k in monthly.columns}

    monthly_summary = monthly.groupby("month").agg(existing_agg)
    monthly_summary.columns = [
        "_".join(col).strip() if isinstance(col, tuple) else col
        for col in monthly_summary.columns
    ]
    monthly_summary = monthly_summary.reset_index()

    # Weekly summary
    weekly = daily.copy()
    weekly["week_start"] = weekly["local_date"].dt.to_period("W").apply(lambda r: r.start_time)
    weekly_summary = weekly.groupby("week_start").agg(existing_agg)
    weekly_summary.columns = [
        "_".join(col).strip() if isinstance(col, tuple) else col
        for col in weekly_summary.columns
    ]
    weekly_summary = weekly_summary.reset_index()

    # Milestones
    valid_weight = daily.dropna(subset=["weight"]).copy()
    valid_weight["weight_change_from_previous"] = valid_weight["weight"].diff()

    first_row = valid_weight.iloc[0]
    latest_row = valid_weight.iloc[-1]
    low_row = valid_weight.loc[valid_weight["weight"].idxmin()]
    high_row = valid_weight.loc[valid_weight["weight"].idxmax()]
    biggest_drop_row = valid_weight.loc[valid_weight["weight_change_from_previous"].idxmin()]
    biggest_gain_row = valid_weight.loc[valid_weight["weight_change_from_previous"].idxmax()]

    last_30 = valid_weight[valid_weight["local_date"] >= valid_weight["local_date"].max() - pd.Timedelta(days=30)]
    last_90 = valid_weight[valid_weight["local_date"] >= valid_weight["local_date"].max() - pd.Timedelta(days=90)]

    milestones = pd.DataFrame(
        [
            ["First valid weight", first_row["local_date"], first_row["weight"]],
            ["Latest valid weight", latest_row["local_date"], latest_row["weight"]],
            ["Total weight change", "", latest_row["weight"] - first_row["weight"]],
            ["Lowest valid weight", low_row["local_date"], low_row["weight"]],
            ["Highest valid weight", high_row["local_date"], high_row["weight"]],
            ["Biggest drop from previous record", biggest_drop_row["local_date"], biggest_drop_row["weight_change_from_previous"]],
            ["Biggest gain from previous record", biggest_gain_row["local_date"], biggest_gain_row["weight_change_from_previous"]],
            ["30-day weight change", "", last_30.iloc[-1]["weight"] - last_30.iloc[0]["weight"] if len(last_30) > 1 else None],
            ["90-day weight change", "", last_90.iloc[-1]["weight"] - last_90.iloc[0]["weight"] if len(last_90) > 1 else None],
            ["Valid daily records", "", len(valid_weight)],
        ],
        columns=["Metric", "Date", "Value"],
    )

    # Dashboard KPIs
    latest_weight = safe_latest(valid_weight["weight"])
    start_weight = valid_weight["weight"].iloc[0]
    latest_bmi = safe_latest(valid_weight["bmi"]) if "bmi" in valid_weight.columns else None
    latest_body_fat = safe_latest(valid_weight["body_fat"]) if "body_fat" in valid_weight.columns else None
    latest_muscle = safe_latest(valid_weight["muscle"]) if "muscle" in valid_weight.columns else None

    dashboard_rows = [
        ["Metric", "Value"],
        ["First Date", valid_weight["local_date"].min()],
        ["Latest Date", valid_weight["local_date"].max()],
        ["Starting Weight", start_weight],
        ["Latest Weight", latest_weight],
        ["Total Weight Change", latest_weight - start_weight],
        ["Lowest Weight", valid_weight["weight"].min()],
        ["Highest Weight", valid_weight["weight"].max()],
        ["Latest BMI", latest_bmi],
        ["Latest Body Fat %", latest_body_fat],
        ["Latest Muscle", latest_muscle],
        ["Valid Daily Records", len(valid_weight)],
        ["Source Clean File", str(CLEAN_MASTER)],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]

    # Create workbook
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_dashboard = wb.create_sheet("Dashboard")
    ws_trend = wb.create_sheet("Trend_Data")
    ws_body = wb.create_sheet("BodyComp_Data")
    ws_monthly = wb.create_sheet("Monthly_Summary")
    ws_weekly = wb.create_sheet("Weekly_Summary")
    ws_milestones = wb.create_sheet("Milestones")
    ws_notes = wb.create_sheet("Source_Notes")

    # Dashboard
    ws_dashboard["A1"] = "Wyze Scale Trend Analysis"
    ws_dashboard["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws_dashboard["A1"].fill = PatternFill("solid", fgColor=BRAND_FILL)
    ws_dashboard.merge_cells("A1:E1")

    for row_idx, row in enumerate(dashboard_rows, start=3):
        ws_dashboard.cell(row=row_idx, column=1, value=row[0])
        ws_dashboard.cell(row=row_idx, column=2, value=row[1])

    ws_dashboard["A3"].fill = PatternFill("solid", fgColor=BRAND_FILL)
    ws_dashboard["B3"].fill = PatternFill("solid", fgColor=BRAND_FILL)
    ws_dashboard["A3"].font = Font(bold=True, color=WHITE)
    ws_dashboard["B3"].font = Font(bold=True, color=WHITE)

    for row in range(4, 17):
        ws_dashboard.cell(row=row, column=1).font = Font(bold=True)
        ws_dashboard.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT_FILL)

    ws_dashboard.column_dimensions["A"].width = 28
    ws_dashboard.column_dimensions["B"].width = 28
    ws_dashboard.column_dimensions["C"].width = 4
    ws_dashboard.column_dimensions["D"].width = 24
    ws_dashboard.column_dimensions["E"].width = 24

    # Data sheets
    write_df(ws_trend, trend)
    write_df(ws_body, bodycomp)
    write_df(ws_monthly, monthly_summary)
    write_df(ws_weekly, weekly_summary)
    write_df(ws_milestones, milestones)

    notes = pd.DataFrame(
        [
            ["Source clean workbook", str(CLEAN_MASTER)],
            ["Output analysis workbook", str(OUTPUT_ANALYSIS)],
            ["Main chart data", "Daily_Analysis from clean master workbook"],
            ["Weight trend rule", "Uses one valid weigh-in per day"],
            ["Body composition rule", "Blank/missing body-composition values are not forced to zero"],
            ["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ],
        columns=["Item", "Description"],
    )
    write_df(ws_notes, notes)

    # Charts on dashboard using hidden-ish data sheet references
    add_line_chart(ws_trend, "Weight Trend", data_col=2, date_col=1, anchor="L2", height=9, width=18)
    add_line_chart(ws_trend, "Weight 30-Day Average", data_col=11, date_col=1, anchor="L20", height=9, width=18)

    if "body_fat" in trend.columns:
        add_line_chart(ws_trend, "Body Fat Trend", data_col=list(trend.columns).index("body_fat") + 1, date_col=1, anchor="L38", height=9, width=18)

    if "muscle" in trend.columns:
        add_line_chart(ws_trend, "Muscle Trend", data_col=list(trend.columns).index("muscle") + 1, date_col=1, anchor="L56", height=9, width=18)

    # Dashboard charts need actual data on dashboard or copied chart objects are not simple.
    # So we add compact dashboard charts directly from Trend_Data by placing them on Dashboard.
    chart = LineChart()
    chart.title = "Weight Over Time"
    chart.y_axis.title = "Weight"
    chart.x_axis.title = "Date"
    chart.height = 9
    chart.width = 18
    data = Reference(ws_trend, min_col=2, min_row=1, max_row=ws_trend.max_row)
    cats = Reference(ws_trend, min_col=1, min_row=2, max_row=ws_trend.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_dashboard.add_chart(chart, "D3")

    chart2 = LineChart()
    chart2.title = "Weight Rolling Averages"
    chart2.y_axis.title = "Weight"
    chart2.x_axis.title = "Date"
    chart2.height = 9
    chart2.width = 18
    data2 = Reference(ws_trend, min_col=10, max_col=11, min_row=1, max_row=ws_trend.max_row)
    cats2 = Reference(ws_trend, min_col=1, min_row=2, max_row=ws_trend.max_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws_dashboard.add_chart(chart2, "D21")

    if "body_fat" in trend.columns and "muscle" in trend.columns:
        chart3 = LineChart()
        chart3.title = "Body Fat and Muscle"
        chart3.y_axis.title = "Value"
        chart3.x_axis.title = "Date"
        chart3.height = 9
        chart3.width = 18
        bf_col = list(trend.columns).index("body_fat") + 1
        muscle_col = list(trend.columns).index("muscle") + 1
        data3 = Reference(ws_trend, min_col=bf_col, max_col=muscle_col, min_row=1, max_row=ws_trend.max_row)
        cats3 = Reference(ws_trend, min_col=1, min_row=2, max_row=ws_trend.max_row)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)
        ws_dashboard.add_chart(chart3, "D39")

    if "weight_mean" in monthly_summary.columns:
        add_bar_chart(
            ws_monthly,
            "Average Weight by Month",
            data_col=list(monthly_summary.columns).index("weight_mean") + 1,
            category_col=1,
            anchor="H2",
            height=9,
            width=18,
        )

    # Number formats
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, float):
                    cell.number_format = "0.00"

    # Dates from pandas may be Timestamp; force date format by likely columns
    for ws in [ws_trend, ws_body, ws_weekly, ws_milestones]:
        for cell in ws["A"]:
            cell.number_format = "yyyy-mm-dd"

    # View niceties
    ws_dashboard.sheet_view.showGridLines = False
    ws_dashboard.freeze_panes = "A3"

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    wb.save(OUTPUT_ANALYSIS)

    print("Done.")
    print(f"Read clean master: {CLEAN_MASTER}")
    print(f"Created analysis workbook: {OUTPUT_ANALYSIS}")
    print("")
    print("Sheets created:")
    for ws in wb.worksheets:
        print(f"- {ws.title}")


if __name__ == "__main__":
    main()
